import re
from docx import Document
import xlrd
from openpyxl import load_workbook
from xlutils.copy import copy


#currently set to replace all instances, need to add counter to replace certain number of instances
def docx_replace_regex(doc_obj, regex, replace):
    for p in doc_obj.paragraphs:
        if regex.search(p.text):
            inline = p.runs
            for i in range(len(inline)):
                if regex.search(inline[i].text):
                    text = regex.sub(replace, inline[i].text)
                    inline[i].text = text
                    #break 
    for table in doc_obj.tables:
        for row in table.rows:
            for cell in row.cells:
                docx_replace_regex(cell, regex , replace)

                
                
# *****************for multiple parts
totalParts = 26
# ****************total number of things to change
variables = 2
filePath = "I:\\Quality\\ECO\\ECO by Year\\ECO 2020\\201102 - Quick Core Auto\\"

# list of parts to iterate through
workbook = xlrd.open_workbook("I:\\R&D\\Active Programs\\54.0 Automatic\Final PN.xlsx")
worksheet = workbook.sheet_by_name('IM-QCA')
                              
for i in range(0,totalParts):
    # please format part numbers in column 1 and rev levels in column 2 or udpate the index below
    partNumber = str(worksheet.cell(i+1, 0).value)
    # handles float to string conversion of part numbers
    
    #rev = worksheet.cell(i+1, 1).value
    rev = "00A"
    # *****************new file path and name
    
    newFilename =  filePath + partNumber + " Rev " + rev + " routing.xlsx"
    filename = filePath + "QCArouting.xlsx"
    pn = worksheet.cell(i+1, 0).value
    description = worksheet.cell(i+1, 1).value
    
    rb = xlrd.open_workbook(filename)
    wb = copy(rb)
    ws = wb.get_sheet(0)
    wb = load_workbook(filename) 
    ws = wb.active 
    ws['B6'] = pn
    ws['H6'] = description

    wb.save(newFilename) 
    

    print(newFilename+" saved!")
